"""
delete_naver_by_excel.py — 네이버 클린프로그램 위반 엑셀 기반 일괄 완전삭제.

엑셀 헤더 (행2): 신고일 | 위반사유 | 상품명 | 몰PID | nv_mid | 비고
- 몰PID = channelProductNo (※ originProductNo 아님)
- nv_mid = 네이버쇼핑 mid (참고용)

처리:
1. 몰PID(channelProductNo) → POST /v1/products/search 로 originProductNo 매핑
2. DELETE /v2/products/origin-products/{originProductNo}
3. listings_pa.status='excluded' (channel='smartstore', channel_product_id=originProductNo)

사용:
    python3 -m backend.purchase.scripts.delete_naver_by_excel <xlsx_path> [--dry-run] [--limit N]
"""
import argparse
import logging
import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parents[3] / ".env")

import openpyxl
import requests

from backend.purchase.services.naver_commerce_service import (
    BASE,
    _gate,
    _get_token,
    delete_product,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)


def load_violations(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[2:]:
        if not r or not r[3]:
            continue
        out.append({
            "report_date": str(r[0]) if r[0] else "",
            "reason": str(r[1]) if r[1] else "",
            "name": str(r[2]) if r[2] else "",
            "origin_product_no": str(r[3]),
            "channel_product_no": str(r[4]) if r[4] else "",
            "note": str(r[5]) if r[5] else "",
        })
    return out


def map_channel_to_origin(channel_nos: list[str]) -> dict[str, str]:
    """몰PID(channelProductNo) → originProductNo 매핑.

    POST /v1/products/search batch (50건씩). 미존재(이미 삭제) 키는 매핑에서 빠짐.
    """
    token = _get_token()
    if not token:
        raise RuntimeError("네이버 토큰 발급 실패")
    mapping: dict[str, str] = {}
    BATCH = 50
    for i in range(0, len(channel_nos), BATCH):
        chunk = channel_nos[i : i + BATCH]
        body = {
            "searchKeywordType": "CHANNEL_PRODUCT_NO",
            "channelProductNos": chunk,
            "page": 1,
            "size": BATCH,
        }
        _gate()
        r = requests.post(
            f"{BASE}/v1/products/search",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=body,
            timeout=20,
        )
        if r.status_code != 200:
            logger.error(f"search 실패: {r.status_code} {r.text[:200]}")
            continue
        for content in r.json().get("contents", []):
            origin = str(content.get("originProductNo") or "")
            for cp in content.get("channelProducts") or []:
                ch = str(cp.get("channelProductNo") or "")
                if ch and origin:
                    mapping[ch] = origin
    return mapping


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path", type=Path)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--limit", type=int, default=None)
    args = ap.parse_args()

    targets = load_violations(args.xlsx_path)
    if args.limit:
        targets = targets[: args.limit]
    logger.info(f"엑셀 행: {len(targets)}건 (몰PID = channelProductNo)")

    # Step 1: channelProductNo → originProductNo 매핑
    channel_nos = [t["origin_product_no"] for t in targets]
    logger.info(f"originProductNo 매핑 시작 (POST /v1/products/search batch)…")
    mapping = map_channel_to_origin(channel_nos)
    logger.info(f"매핑 완료: {len(mapping)}/{len(channel_nos)}건")

    missing = [c for c in channel_nos if c not in mapping]
    if missing:
        logger.warning(f"매핑 미발견 (이미 삭제 가능): {len(missing)}건 — 예: {missing[:5]}")

    if args.dry_run:
        for t in targets[:5]:
            ch = t["origin_product_no"]
            origin = mapping.get(ch, "<not found>")
            logger.info(f"  [DRY] channel={ch} → origin={origin} | {t['name'][:40]}")
        logger.info(f"  ... 총 {len(targets)}건, 매핑 {len(mapping)}건 (dry-run)")
        return

    db = Path(__file__).resolve().parents[1] / "purchase.db"
    conn = sqlite3.connect(str(db))

    ok = 0
    fail = 0
    skipped = 0
    fail_msgs = []
    delete_targets = [(t, mapping[t["origin_product_no"]]) for t in targets if t["origin_product_no"] in mapping]
    skipped = len(targets) - len(delete_targets)

    for i, (t, origin_no) in enumerate(delete_targets):
        success, err = delete_product(origin_no)
        if success:
            ok += 1
            conn.execute(
                """UPDATE listings_pa
                   SET status='excluded',
                       error_message='네이버 클린프로그램 중복상품 신고 — 완전삭제 (20260427)',
                       last_synced_at=datetime('now')
                   WHERE channel='smartstore' AND channel_product_id=?""",
                (origin_no,),
            )
            conn.commit()
        else:
            fail += 1
            if len(fail_msgs) < 20:
                fail_msgs.append((f"channel={t['origin_product_no']} origin={origin_no}", err))

        if (i + 1) % 5 == 0 or (i + 1) == len(delete_targets):
            logger.info(f"  progress {i+1}/{len(delete_targets)} ok={ok} fail={fail}")

    logger.info(f"완료: 삭제 성공 {ok}, 실패 {fail}, 매핑 미발견 skip {skipped}")
    for label, err in fail_msgs:
        logger.warning(f"  {label}: {err}")


if __name__ == "__main__":
    main()
